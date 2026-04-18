from flask import Flask, render_template, request, redirect, url_for, jsonify, send_file, flash, session
import sqlite3, os, json, csv, io, requests
from datetime import datetime, date
from werkzeug.utils import secure_filename
from werkzeug.security import generate_password_hash, check_password_hash
from functools import wraps
import openpyxl

app = Flask(__name__)
app.secret_key = 'sreports-secret-2025'
UPLOAD_FOLDER = 'uploads'
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
DB = 'sreports.db'
OLLAMA_URL = 'http://localhost:11434/api/generate'

# ─── DB ───────────────────────────────────────────────────────────────────────

def get_db():
    conn = sqlite3.connect(DB)
    conn.row_factory = sqlite3.Row
    return conn

def init_db():
    with get_db() as conn:
        conn.executescript('''
            CREATE TABLE IF NOT EXISTS users (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL,
                email TEXT UNIQUE NOT NULL,
                password TEXT NOT NULL,
                role TEXT DEFAULT 'user',
                created_at TEXT DEFAULT (datetime('now'))
            );
            CREATE TABLE IF NOT EXISTS companies (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL,
                sector TEXT,
                created_at TEXT DEFAULT (datetime('now'))
            );
            CREATE TABLE IF NOT EXISTS reports (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                company_id INTEGER,
                user_id INTEGER,
                title TEXT NOT NULL,
                report_type TEXT NOT NULL,
                period TEXT NOT NULL,
                period_label TEXT,
                status TEXT DEFAULT 'draft',
                data_json TEXT,
                analysis TEXT,
                created_at TEXT DEFAULT (datetime('now')),
                FOREIGN KEY (company_id) REFERENCES companies(id),
                FOREIGN KEY (user_id) REFERENCES users(id)
            );
            CREATE TABLE IF NOT EXISTS kpi_entries (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                report_id INTEGER,
                category TEXT,
                metric TEXT,
                value REAL,
                unit TEXT,
                period TEXT,
                FOREIGN KEY (report_id) REFERENCES reports(id)
            );
        ''')
        cur = conn.execute("SELECT COUNT(*) FROM companies")
        if cur.fetchone()[0] == 0:
            conn.execute("INSERT INTO companies (name, sector) VALUES ('Empresa Demo S.L.', 'Tecnología')")

init_db()

# ─── AUTH ─────────────────────────────────────────────────────────────────────

def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated

@app.route('/login', methods=['GET', 'POST'])
@app.route('/login', methods=['GET', 'POST'])
def login():
    if 'user_id' in session:
        return redirect(url_for('index'))
    error = None
    if request.method == 'POST':
        email = request.form.get('email', '').strip().lower()
        password = request.form.get('password', '')
        db = get_db()
        user = db.execute("SELECT * FROM users WHERE email=?", (email,)).fetchone()
        if user and check_password_hash(user['password'], password):
            session['user_id'] = user['id']
            session['user_name'] = user['name']
            session['user_email'] = user['email']
            return redirect(url_for('index'))
        error = 'Email o contraseña incorrectos.'
    return render_template('login.html', error=error)

@app.route('/register', methods=['GET', 'POST'])
@app.route('/register', methods=['GET', 'POST'])
def register():
    if 'user_id' in session:
        return redirect(url_for('index'))
    error = None
    if request.method == 'POST':
        name = request.form.get('name', '').strip()
        email = request.form.get('email', '').strip().lower()
        password = request.form.get('password', '')
        confirm = request.form.get('confirm', '')
        if not name or not email or not password:
            error = 'Completa todos los campos.'
        elif password != confirm:
            error = 'Las contraseñas no coinciden.'
        elif len(password) < 6:
            error = 'La contraseña debe tener al menos 6 caracteres.'
        else:
            db = get_db()
            existing = db.execute("SELECT id FROM users WHERE email=?", (email,)).fetchone()
            if existing:
                error = 'Este email ya está registrado.'
            else:
                hashed = generate_password_hash(password)
                cur = db.execute("INSERT INTO users (name, email, password) VALUES (?,?,?)", (name, email, hashed))
                db.commit()
                session['user_id'] = cur.lastrowid
                session['user_name'] = name
                session['user_email'] = email
                return redirect(url_for('index'))
    return render_template('register.html', error=error)

@app.route('/logout')
@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))

# ─── HELPERS ──────────────────────────────────────────────────────────────────

REPORT_TYPES = {
    'ventas': {'label': 'Ventas / KPIs', 'icon': '📈', 'color': '#E8186D'},
    'financiero': {'label': 'Financiero', 'icon': '💰', 'color': '#a855f7'},
    'rrhh': {'label': 'RRHH / Equipo', 'icon': '👥', 'color': '#3b82f6'},
    'marketing': {'label': 'Marketing', 'icon': '📣', 'color': '#10b981'},
}

def parse_uploaded_file(file):
    filename = secure_filename(file.filename)
    ext = filename.rsplit('.', 1)[-1].lower()
    rows = []
    if ext == 'csv':
        content = file.read().decode('utf-8-sig')
        reader = csv.DictReader(io.StringIO(content))
        rows = list(reader)
    elif ext in ['xlsx', 'xls']:
        wb = openpyxl.load_workbook(file, data_only=True)
        ws = wb.active
        headers = [str(cell.value) for cell in next(ws.iter_rows(max_row=1))]
        for row in ws.iter_rows(min_row=2, values_only=True):
            rows.append(dict(zip(headers, [str(v) if v is not None else '' for v in row])))
    return rows

def call_ollama(prompt):
    try:
        r = requests.post(OLLAMA_URL, json={
            'model': 'llama3',
            'prompt': prompt,
            'stream': False
        }, timeout=120)
        return r.json().get('response', 'Sin respuesta del modelo.')
    except Exception as e:
        return f'[Ollama no disponible: {e}]'

def generate_analysis(report_type, data_json, period_label):
    data = json.loads(data_json) if isinstance(data_json, str) else data_json
    prompt = f"""Eres un analista de negocio experto. Analiza los siguientes datos empresariales y genera un informe ejecutivo profesional en español.

Tipo de informe: {REPORT_TYPES.get(report_type, {}).get('label', report_type)}
Período: {period_label}
Datos: {json.dumps(data, ensure_ascii=False, indent=2)}

Estructura tu respuesta así:
1. RESUMEN EJECUTIVO (2-3 frases clave)
2. ANÁLISIS DETALLADO (hallazgos principales)
3. PUNTOS CRÍTICOS (máx. 3 alertas o riesgos)
4. RECOMENDACIONES (máx. 3 acciones concretas)

Sé directo, profesional y orientado a datos. No uses emojis."""
    return call_ollama(prompt)

# ─── ROUTES ───────────────────────────────────────────────────────────────────

@app.route('/')
@login_required
def index():
    db = get_db()
    companies = db.execute("SELECT * FROM companies ORDER BY name").fetchall()
    reports = db.execute("""
        SELECT r.*, c.name as company_name 
        FROM reports r LEFT JOIN companies c ON r.company_id = c.id
        ORDER BY r.created_at DESC LIMIT 20
    """).fetchall()
    stats = {
        'total_reports': db.execute("SELECT COUNT(*) FROM reports").fetchone()[0],
        'this_month': db.execute("SELECT COUNT(*) FROM reports WHERE strftime('%Y-%m', created_at) = strftime('%Y-%m', 'now')").fetchone()[0],
        'companies': db.execute("SELECT COUNT(*) FROM companies").fetchone()[0],
        'pending': db.execute("SELECT COUNT(*) FROM reports WHERE status='draft'").fetchone()[0],
    }
    return render_template('index.html', companies=companies, reports=reports, stats=stats, report_types=REPORT_TYPES)

@app.route('/report/new', methods=['GET', 'POST'])
@login_required
def new_report():
    db = get_db()
    companies = db.execute("SELECT * FROM companies ORDER BY name").fetchall()

    if request.method == 'POST':
        company_id = request.form.get('company_id')
        title = request.form.get('title')
        report_type = request.form.get('report_type')
        period = request.form.get('period')
        period_label = request.form.get('period_label', period)
        data_source = request.form.get('data_source')
        data = {}

        if data_source == 'file' and 'data_file' in request.files:
            f = request.files['data_file']
            if f.filename:
                rows = parse_uploaded_file(f)
                data = {'source': 'file', 'rows': rows}
        elif data_source == 'manual':
            metrics = request.form.getlist('metric_name[]')
            values = request.form.getlist('metric_value[]')
            units = request.form.getlist('metric_unit[]')
            data = {'source': 'manual', 'metrics': [
                {'name': m, 'value': v, 'unit': u}
                for m, v, u in zip(metrics, values, units) if m
            ]}
        elif data_source == 'form':
            data = {'source': 'form', 'fields': dict(request.form)}

        data_json = json.dumps(data, ensure_ascii=False)
        analysis = generate_analysis(report_type, data_json, period_label)

        cur = db.execute("""
            INSERT INTO reports (company_id, title, report_type, period, period_label, status, data_json, analysis)
            VALUES (?,?,?,?,?,?,?,?)
        """, (company_id, title, report_type, period, period_label, 'published', data_json, analysis))
        db.commit()
        report_id = cur.lastrowid
        return redirect(url_for('view_report', report_id=report_id))

    return render_template('new_report.html', companies=companies, report_types=REPORT_TYPES)

@app.route('/report/<int:report_id>')
@login_required
def view_report(report_id):
    db = get_db()
    report = db.execute("""
        SELECT r.*, c.name as company_name, c.sector
        FROM reports r LEFT JOIN companies c ON r.company_id = c.id
        WHERE r.id = ?
    """, (report_id,)).fetchone()
    if not report:
        return redirect(url_for('index'))
    data = json.loads(report['data_json']) if report['data_json'] else {}
    return render_template('report.html', report=report, data=data, report_types=REPORT_TYPES)

@app.route('/report/<int:report_id>/pdf')
@login_required
def export_pdf(report_id):
    db = get_db()
    report = db.execute("""
        SELECT r.*, c.name as company_name, c.sector
        FROM reports r LEFT JOIN companies c ON r.company_id = c.id
        WHERE r.id = ?
    """, (report_id,)).fetchone()
    data = json.loads(report['data_json']) if report['data_json'] else {}
    html_content = render_template('report_pdf.html', report=report, data=data, report_types=REPORT_TYPES)
    
    try:
        from weasyprint import HTML
        pdf_bytes = HTML(string=html_content).write_pdf()
        return send_file(
            io.BytesIO(pdf_bytes),
            mimetype='application/pdf',
            as_attachment=True,
            download_name=f'informe_{report_id}_{report["period"]}.pdf'
        )
    except ImportError:
        return html_content, 200, {'Content-Type': 'text/html'}

@app.route('/report/<int:report_id>/regenerate', methods=['POST'])
@login_required
def regenerate_analysis(report_id):
    db = get_db()
    report = db.execute("SELECT * FROM reports WHERE id=?", (report_id,)).fetchone()
    analysis = generate_analysis(report['report_type'], report['data_json'], report['period_label'])
    db.execute("UPDATE reports SET analysis=? WHERE id=?", (analysis, report_id))
    db.commit()
    return jsonify({'analysis': analysis})

@app.route('/company/add', methods=['POST'])
@login_required
def add_company():
    name = request.form.get('name')
    sector = request.form.get('sector', '')
    if name:
        db = get_db()
        db.execute("INSERT INTO companies (name, sector) VALUES (?,?)", (name, sector))
        db.commit()
    return redirect(url_for('index'))

@app.route('/report/<int:report_id>/delete', methods=['POST'])
@login_required
def delete_report(report_id):
    db = get_db()
    db.execute("DELETE FROM reports WHERE id=?", (report_id,))
    db.commit()
    return redirect(url_for('index'))

@app.route('/api/reports')
@login_required
def api_reports():
    db = get_db()
    reports = db.execute("""
        SELECT r.id, r.title, r.report_type, r.period_label, r.status, r.created_at, c.name as company
        FROM reports r LEFT JOIN companies c ON r.company_id = c.id
        ORDER BY r.created_at DESC
    """).fetchall()
    return jsonify([dict(r) for r in reports])

if __name__ == '__main__':
    app.run(debug=True, port=5000)
